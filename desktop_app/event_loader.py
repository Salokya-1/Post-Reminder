from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


@dataclass(frozen=True)
class Event:
    name: str
    date: date
    notes: str = ""


def _normalize_header(value: object) -> str:
    return str(value or "").strip().lower()


def _parse_date(value: object) -> date | None:
    if value is None:
        return None

    if isinstance(value, date) and not isinstance(value, datetime):
        return value

    if isinstance(value, datetime):
        return value.date()

    text = str(value).strip()
    if not text:
        return None

    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%b %d, %Y", "%B %d, %Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    return None


def load_events_from_csv(path: Path) -> list[Event]:
    import csv

    events: list[Event] = []
    with path.open("r", encoding="utf-8") as file:
        reader = csv.DictReader(file)
        for row in reader:
            name = (row.get("name") or row.get("event") or "").strip()
            notes = (row.get("notes") or row.get("caption") or "").strip()
            event_date = _parse_date(row.get("date"))

            if not name or event_date is None:
                continue

            events.append(Event(name=name, date=event_date, notes=notes))

    return _dedupe(events)


def load_events_from_xlsx(path: Path) -> list[Event]:
    workbook = load_workbook(path, data_only=True)
    events: list[Event] = []

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [_normalize_header(cell) for cell in rows[0]]

        date_idx = _find_col(headers, ("date", "day"))
        name_idx = _find_col(headers, ("name", "event", "festival", "post", "title"))
        notes_idx = _find_col(headers, ("notes", "caption", "description", "details", "content"))

        for row in rows[1:]:
            event_date = _parse_date(row[date_idx] if date_idx is not None and date_idx < len(row) else None)

            if name_idx is not None and name_idx < len(row):
                name = str(row[name_idx] or "").strip()
            else:
                name = _first_text_cell(row, skip_index=date_idx)

            notes = ""
            if notes_idx is not None and notes_idx < len(row):
                notes = str(row[notes_idx] or "").strip()

            if event_date is None or not name:
                continue

            events.append(Event(name=name, date=event_date, notes=notes))

    workbook.close()
    return _dedupe(events)


def _find_col(headers: list[str], keywords: tuple[str, ...]) -> int | None:
    for idx, header in enumerate(headers):
        for keyword in keywords:
            if keyword in header:
                return idx
    return None


def _first_text_cell(row: tuple[object, ...], skip_index: int | None) -> str:
    for idx, value in enumerate(row):
        if skip_index is not None and idx == skip_index:
            continue

        text = str(value or "").strip()
        if text:
            parsed = _parse_date(text)
            if parsed is None:
                return text
    return ""


def _dedupe(events: list[Event]) -> list[Event]:
    seen: set[tuple[str, date]] = set()
    result: list[Event] = []

    for event in sorted(events, key=lambda item: item.date):
        key = (event.name.lower(), event.date)
        if key in seen:
            continue
        seen.add(key)
        result.append(event)

    return result
